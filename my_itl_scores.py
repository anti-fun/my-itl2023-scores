#!/usr/bin/env python3
import pandas as pd
import numpy as np
import openpyxl as px
import requests
from datetime import datetime
from math import log, isnan

ENTRANT_ID = 0 # your entrant id here
ADDITIONAL_RIVAL_IDS = []

def current_datetime():
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")

def get_entrant_data(eid):
    return requests.get(f"https://itl2023.groovestats.com/api/entrant/{eid}").json()["data"]

def get_rival_ids(eid):
    ladder = requests.get(f"https://itl2023.groovestats.com/api/entrant/{eid}/stats").json()["data"]["ladder"]
    return [entry["id"] for entry in ladder if entry["type"] == "rival"]

def load_entrant_data(data):
    return pd.DataFrame.from_records(
        data=data["topScores"],
        index=[score["chartHash"] for score in data["topScores"]],
    )

def points_pct(ex):
    return round(log(min((ex), 50)+1, 1.1032889141348) + 61**(max(0, (ex)-50)/50)-1, 2)

def calculate_points(ex, max_points):
    return int(max_points * points_pct(ex) / 100)

# brute force inverting the points function
preimage = [x/100 for x in range(10001)]
image = np.array([points_pct(x) for x in preimage])
def points_pct_inverse (pct):
    idx = min((np.abs(image - pct)).argmin() + 1, 10000)
    return preimage[idx]

def ex_from_points(points, max_points):
    return float(points_pct_inverse(100 * points / max_points))

def ex_to_gain_rp(chartHash):
    pts, ex, pts_max = top_scores.loc[chartHash, ["points", "ex", "points max"]]

    if pts >= rp_thresh:
        return ex + .01
    if pts_max <= rp_thresh:
        return None
    return ex_from_points(rp_thresh + 1, pts_max)

if not ENTRANT_ID:
    print("Need to fill in ENTRANT_ID. Exiting.")
    exit()

# get initial entrant data and load it
entrant_data = get_entrant_data(ENTRANT_ID)
entrant_name = entrant_data["entrant"]["name"]

top_scores = load_entrant_data(entrant_data).sort_values(by="points", ascending=False)
top_scores["rank"] = range(1, len(top_scores.index)+1)
top_scores["ex"] /= 100
for field in ["lastUpdated", "lastImproved", "dateAdded"]:
    top_scores[field] = top_scores[field].apply(pd.Timestamp)
    top_scores[field] = top_scores[field].dt.tz_localize(None)

charts = pd.DataFrame.from_records(
    data=entrant_data["charts"],
    exclude=["id", "hash"],
    index=[chart["hash"] for chart in entrant_data["charts"]],
)

# get rival data and load it
rival_ids = get_rival_ids(ENTRANT_ID) + ADDITIONAL_RIVAL_IDS
rival_col_names = []
for rival_id in rival_ids:
    r_data = get_entrant_data(rival_id)
    r_suffix = f' ({r_data["entrant"]["name"]})'
    r_diff_col_name = f"ex diff{r_suffix}"

    # show charts that rivals have unlocked even if you haven't
    r_charts = pd.DataFrame.from_records(
        data=r_data["charts"],
        exclude=["id", "hash"],
        index=[chart["hash"] for chart in r_data["charts"]],
    )
    charts = pd.concat([charts, r_charts]).drop_duplicates()

    r_df = load_entrant_data(r_data)[["ex", "clearType"]]
    r_df["ex"] /= 100

    r_df[r_diff_col_name] = top_scores["ex"] - r_df["ex"]
    top_scores = top_scores.join(r_df, how="outer", rsuffix=r_suffix, validate="1:1")

    # save column names for later reorganizing
    rival_col_names.append(f"ex{r_suffix}")
    rival_col_names.append(r_diff_col_name)
    rival_col_names.append(f"clearType{r_suffix}")

# force romanized titles
for chartHash in charts.index:
    if charts.loc[chartHash, "titleRomaji"]:
        charts.loc[chartHash, "title"] = charts.loc[chartHash, "titleRomaji"]
    if charts.loc[chartHash, "subtitleRomaji"]:
        charts.loc[chartHash, "subtitle"] = charts.loc[chartHash, "subtitleRomaji"]
charts.drop(columns=["titleRomaji", "subtitleRomaji"], inplace=True)

# add chart info into scores DF
top_scores = top_scores.join(charts, how="right", rsuffix=" max", validate="1:1").sort_values(by="points", ascending=False)
top_scores["points percent"] = round(100 * top_scores["points"] / top_scores["points max"], 2)

# add "upscore to gain RP" columns
rp_thresh = 0 if isnan(top_scores.iloc[74]["points"]) else top_scores.iloc[74]["points"]
top_scores["ex to gain"] = [ex_to_gain_rp(x) for x in top_scores.index]
top_scores["upscore to gain"] = top_scores["ex to gain"] - top_scores["ex"]

# move the interesting stuff to the beginning and the uninteresting stuff to the end
interesting_cols = [
    "title",
    "meter",
    "ex",
    "points",
    "points max",
    "points percent",
    "rank",
    "clearType",
    "ex to gain",
    "upscore to gain",
] + rival_col_names
uninteresting_cols = [
    "id",
    "entrantId",
    "dateAdded",
    "chartHash",
    "hashOriginal",
]

new_cols = \
    interesting_cols + \
    [x for x in top_scores.columns if x not in interesting_cols+uninteresting_cols] + \
    uninteresting_cols

top_scores = top_scores[new_cols]

# create spreadsheet
dt = current_datetime()
fn = f"scores_{ENTRANT_ID}-{entrant_name}_{dt}"
top_scores.to_excel(fn + ".xlsx", index=False, freeze_panes=[1,1])
wb = px.load_workbook(fn + ".xlsx")
ws = wb.active

# resize columns
ws.auto_filter.ref = ws.dimensions
for column_cells in ws.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (px.utils.get_column_letter(column_cells[0].column))
    if new_column_length > 0:
        ws.column_dimensions[new_column_letter].width = max(new_column_length + 4.6, 11)

# add conditional formatting
# rival ex difference gradient
grad_rule = px.formatting.rule.ColorScaleRule(
    start_type='percentile', start_value=1, start_color='F8696B',
    mid_type='num', mid_value=0, mid_color='FFFFFF',
    end_type='percentile', end_value=100, end_color='63BE7B',
)
grey_fill = px.styles.PatternFill(
    start_color='D9D9D9',
    end_color='D9D9D9',
    fill_type='solid',
)
ex_diff_col_nums = [list(top_scores.columns).index(x)+1 for x in top_scores.columns if 'ex diff' in x]
for i in ex_diff_col_nums:
    col = px.utils.get_column_letter(i)
    range_str = f'{col}2:{col}{len(top_scores.index)+1}'
    ws.conditional_formatting.add(range_str, grad_rule)
    ws.conditional_formatting.add(
        range_str, 
        px.formatting.rule.FormulaRule(formula=[f'ISBLANK({col}2)'], stopIfTrue=True, fill=grey_fill)
    )

# combo lamp highlighting
clear_type_fills = [
    # fc
    px.styles.PatternFill(
        start_color='92D050',
        end_color='92D050',
        fill_type='solid',
    ),
    # fec
    px.styles.PatternFill(
        start_color='FFC000',
        end_color='FFC000',
        fill_type='solid',
    ),
    # quad
    px.styles.PatternFill(
        start_color='00B0F0',
        end_color='00B0F0',
        fill_type='solid',
    ),
    # quint
    px.styles.PatternFill(
        start_color='D02085',
        end_color='D02085',
        fill_type='solid',
    ),
]
clear_type_col_nums = [list(top_scores.columns).index(x)+1 for x in top_scores.columns if 'clearType' in x]
for i in clear_type_col_nums:
    col = px.utils.get_column_letter(i)
    range_str = f'{col}2:{col}{len(top_scores.index)+1}'

    ct = 2
    for fill in clear_type_fills:
        ws.conditional_formatting.add(
            range_str, 
            px.formatting.rule.FormulaRule(formula=[f'{col}2={ct}'], stopIfTrue=True, fill=fill)
        )
        ct += 1

wb.save(fn + ".xlsx")